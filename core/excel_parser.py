import random
import logging
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

def parse_excel_to_tasks(filepath):
    """
    Parses the specific Excel file and returns a list of tasks.
    Column B: Base prompt
    Column C: Target image count (total)
    Column D: Output path
    Columns E to M: Feature prompts
    """
    try:
        wb = load_workbook(filepath, data_only=True)
        sheet = wb.active
    except Exception as e:
        logger.error(f"Failed to load excel file {filepath}: {e}")
        raise e

    # Extract all features from Columns E (5) to the last column
    # The structure is a list of lists: [[features for col E], [features for col F], ...]
    feature_columns = []
    for col_idx in range(5, sheet.max_column + 1):
        features = []
        for row_idx in range(2, sheet.max_row + 1):
            val = sheet.cell(row=row_idx, column=col_idx).value
            if val is not None and str(val).strip() != "":
                features.append(str(val).strip())
        # Only add the column to our feature lists if it actually has options
        if features:
            feature_columns.append(features)

    tasks = []

    # Iterate over rows for Base Prompts (Column B = 2)
    for row_idx in range(2, sheet.max_row + 1):
        base_prompt_val = sheet.cell(row=row_idx, column=2).value
        
        # If Column B has a value, we process this as a base prompt task
        if base_prompt_val is not None and str(base_prompt_val).strip() != "":
            base_prompt = str(base_prompt_val).strip()
            
            # Target count from Column C (3)
            target_count_val = sheet.cell(row=row_idx, column=3).value
            try:
                target_count = int(target_count_val) if target_count_val else 1
            except ValueError:
                target_count = 1
                
            # Output path from Column D (4)
            output_path_val = sheet.cell(row=row_idx, column=4).value
            output_path = str(output_path_val).strip() if output_path_val else None

            # Generate target_count individual tasks with randomly sampled features
            for _ in range(target_count):
                sampled_features = []
                for features in feature_columns:
                    if features:
                        sampled_features.append(random.choice(features))
                
                # Combine base prompt and selected features
                combined_prompt = base_prompt
                if sampled_features:
                    combined_prompt += " " + " ".join(sampled_features)

                tasks.append({
                    "prompt": combined_prompt,
                    "target_count": 1,
                    "output_path": output_path
                })
                
    return tasks
